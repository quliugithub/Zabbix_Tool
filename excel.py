from __future__ import annotations

from typing import List
from pathlib import Path
from openpyxl import load_workbook
from models import InstallRequest


def _coerce_numeric(val):
    try:
        return int(val)
    except Exception:
        return val


def parse_excel(path: str, column_map: dict | None = None) -> List[InstallRequest]:
    """
    Parse Excel file into InstallRequest list.
    column_map: mapping from logical fields to column names, defaults for: hostname, ip, os_type, env, port, ssh_user, ssh_password, ssh_port, visible_name, jmx_port, template_id, group_id, note
    """
    alias = {
        "hostname": ["hostname", "host", "主机名", "主机", "主机名称"],
        "ip": ["ip", "IP", "主机ip", "ip地址", "地址"],
        "os_type": ["os_type", "os", "系统", "系统类型"],
        "env": ["env", "环境", "环境标识"],
        "port": ["port", "agent_port", "agent端口", "agent 端口", "agent端口号"],
        "ssh_user": ["ssh_user", "ssh用户", "ssh user"],
        "ssh_password": ["ssh_password", "ssh密码", "ssh pass", "密码"],
        "ssh_port": ["ssh_port", "ssh端口", "ssh port"],
        "visible_name": ["visible_name", "可见名称", "显示名", "别名"],
        "jmx_port": ["jmx_port", "jmx端口", "jmx port"],
        "template_ids": ["template_ids", "template_id", "模板", "模板id", "模板ids"],
        "group_ids": ["group_ids", "group_id", "群组", "群组id", "组id"],
        "proxy_id": ["proxy_id", "代理", "代理id", "proxy"],
        "web_monitor_url": ["web_monitor_url", "web监控url", "web监控", "web监控地址", "web url", "web"],
        "note": ["note", "备注"],
    }
    if column_map:
        for k, v in column_map.items():
            alias[k] = [v]

    ext = Path(path).suffix.lower()
    allowed_ext = {".xlsx", ".xlsm", ".xltx", ".xltm"}
    if ext not in allowed_ext:
        raise ValueError(f"请上传 Excel 文件 ({', '.join(sorted(allowed_ext))})，当前文件类型: {ext or '未知'}")

    rows = []
    wb = load_workbook(path)
    ws = wb.active
    headers = {
        str(cell.value).strip().lower(): idx
        for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=0)
        if cell.value
    }
    for row in ws.iter_rows(min_row=2):
        entry = {}
        for field, names in alias.items():
            col = next((n for n in names if n.lower() in headers), None)
            if col:
                cell = row[headers[col.lower()]]
                entry[field] = cell.value
        rows.append(entry)

    requests: List[InstallRequest] = []
    for data in rows:
        # 统一去除空字符串
        for k, v in list(data.items()):
            if isinstance(v, str):
                data[k] = v.strip()
            if data[k] == "":
                data[k] = None

        if not data.get("ip"):
            continue
        # 默认 os_type
        data["os_type"] = data.get("os_type") or data.get("os") or "linux"
        # 强制把数字字段转为 int
        for key in ("port", "ssh_port", "jmx_port"):
            if key in data and data[key] not in (None, ""):
                data[key] = _coerce_numeric(data[key])
        # 多选字段按逗号拆分
        for key in ("template_ids", "group_ids"):
            if isinstance(data.get(key), str):
                parts = [p.strip() for p in data[key].split(",") if p.strip()]
                data[key] = parts or None
        # 将应为字符串的字段强制转为字符串，避免数字导致校验失败
        for key in ("hostname", "visible_name", "ssh_user", "ssh_password", "env", "os_type", "note"):
            if data.get(key) not in (None, ""):
                data[key] = str(data[key])
        req = InstallRequest(**data)
        # 转存为 dict 供前端展示时避免 IPv4Address 无法 JSON 序列化
        requests.append(req)
    return requests
